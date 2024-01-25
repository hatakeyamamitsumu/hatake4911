import streamlit as st
from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import PatternFill

def image_to_excel(image_path, output_filename):
    # 画像ファイルを開く
    image = Image.open(image_path)

    # 画像の幅と高さを取得
    width, height = image.size

    # 新しいExcelブックを作成
    workbook = Workbook()
    sheet = workbook.active

    # 画像の各ピクセルをExcelセルに対応させて色を設定
    for y in range(height):
        for x in range(width):
            pixel_color = image.getpixel((x, y))
            hex_color = f'{pixel_color[0]:02X}{pixel_color[1]:02X}{pixel_color[2]:02X}'  # RGBを16進数に変換
            fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')
            cell = sheet.cell(row=y + 1, column=x + 1)
            cell.fill = fill

    # Excelファイルを保存
    workbook.save(output_filename)
    print(f'Image colors saved in Excel: {output_filename}')

# アップローダーを作成
uploaded_file = st.file_uploader('画像ファイルをアップロード', type='png')

# 画像がアップロードされたら処理を実行
if uploaded_file is not None:
    image_path = uploaded_file.name
    output_excel_filename = st.text_input('出力ファイル名')
    image_to_excel(image_path, output_excel_filename)
    st.success('画像の色をExcelシートに保存しました。')

# アップロードされた画像をプレビュー
if uploaded_file is not None:
    image = Image.open(uploaded_file)
    st.image(image)

# ダウンロードリンクを作成
if uploaded_file is not None:
    output_excel_file = f'{output_excel_filename}.xlsx'
    st.download_button(
        label='Excelファイルをダウンロード',
        data=output_excel_file,
        filename=output_excel_filename,
        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
