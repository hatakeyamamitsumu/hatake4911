import streamlit as st
from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import os
import shutil
import tempfile
import numpy as np

def perspective_transform(img, dst_pts):
    src_pts = np.array([[0, 0], [img.shape[1], 0], [img.shape[1], img.shape[0]], [0, img.shape[0]]], dtype=np.float32)
    matrix = cv2.getPerspectiveTransform(src_pts, dst_pts)
    result = cv2.warpPerspective(img, matrix, (img.shape[1], img.shape[0]))
    return result

def swirl_image(image_array, strength=10, radius=1000, center_x=None, center_y=None):
    height, width, _ = image_array.shape

    if center_x is None:
        center_x = width // 2
    if center_y is None:
        center_y = height // 2

    output_array = np.empty_like(image_array)

    for y in range(height):
        for x in range(width):
            dx = x - center_x
            dy = y - center_y
            distance = np.sqrt(dx ** 2 + dy ** 2)
            angle = np.arctan2(dy, dx)

            if distance == 0:
                new_x = x
                new_y = y
            else:
                swirl_strength = strength * (radius / distance) ** 0.2
                new_x = int(center_x + np.cos(angle + swirl_strength) * distance)
                new_y = int(center_y + np.sin(angle + swirl_strength) * distance)

            if 0 <= new_x < width and 0 <= new_y < height:
                output_array[y, x] = image_array[new_y, new_x]
            else:
                output_array[y, x] = 0

    return output_array

def image_to_excel(image, output_filename):
    width, height = image.size

    workbook = Workbook()
    sheet = workbook.active

    for y in range(height):
        for x in range(width):
            pixel_color = image.getpixel((x, y))
            hex_color = f'{pixel_color[0]:02X}{pixel_color[1]:02X}{pixel_color[2]:02X}'
            fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')
            cell = sheet.cell(row=y + 1, column=x + 1)
            cell.fill = fill

    temp_folder = tempfile.mkdtemp()
    temp_filepath = os.path.join(temp_folder, output_filename)
    workbook.save(temp_filepath)
    return temp_filepath

def main():
    st.title("画像処理アプリ")

    option = st.sidebar.selectbox("処理を選択してください", ("ひし形に変形", "渦巻き", "画像をExcelに転記"))

    if option == "ひし形に変形":
        st.subheader("写真をひし形に変形")
        st.write("写真をひし形に変形させるアプリです。")

        uploaded_file = st.file_uploader("写真をアップロードしてください", type=["jpg", "jpeg", "png"])

        if uploaded_file is not None:
            image = cv2.imdecode(np.fromstring(uploaded_file.read(), np.uint8), 1)

            default_values = [0.25, 0.2, 0.75, 0.1, 1.0, 0.5, 0.0, 0.8]
            dst_pts = np.array([[image.shape[1] * default_values[i], image.shape[0] * default_values[i + 1]] for i in range(0, len(default_values), 2)], dtype=np.float32)

            st.subheader("写真の四隅の座標を変更してください:")
            for i in range(4):
                row = st.slider(f"Point {i + 1} X", 0.0, 1.0, default_values[i * 2], 0.01)
                col = st.slider(f"Point {i + 1} Y", 0.0, 1.0, default_values[i * 2 + 1], 0.01)
                dst_pts[i] = [row * image.shape[1], col * image.shape[0]]

            result = perspective_transform(image, dst_pts)

            st.image(image, caption='Original Image', use_column_width=True, channels="BGR")
            st.image(result, caption='Perspective Transformed Image', use_column_width=True, channels="BGR")

            st.download_button(
                label="ダウンロードボタン",
                data=cv2.imencode('.jpg', result)[1].tobytes(),
                file_name='Perspective_Transformed_Image.jpg',
                mime='image/jpg'
            )

    elif option == "渦巻き":
        st.subheader("写真を渦巻きに変換")
        st.write("写真を渦巻き状に変形させるアプリです。")

        uploaded_file = st.file_uploader("画像をアップロードしてください", type=["jpg", "jpeg", "png"])

        if uploaded_file is not None:
            image = Image.open(uploaded_file)
            image_array = np.array(image)

            strength = st.slider("渦の強さ", min_value=1, max_value=20, value=10)
            radius = st.slider("渦の半径サイズ", min_value=100, max_value=2000, value=1000)
            center_x = st.slider("中心のX座標", min_value=-image_array.shape[1], max_value=2*image_array.shape[1], value=image_array.shape[1] // 2)
            center_y = st.slider("中心のY座標", min_value=-image_array.shape[0], max_value=2*image_array.shape[0], value=image_array.shape[0] // 2)
            display_width = st.slider("画像サイズ", min_value=100, max_value=1000, value=600)

            processed_image_array = swirl_image(image_array, strength, radius, center_x, center_y)

            st.image([image, Image.fromarray(processed_image_array)], caption=["Original Image", "Swirled Image"], width=display_width)

            processed_image = Image.fromarray(processed_image_array)
            img_byte_array = io.BytesIO()
            processed_image.save(img_byte_array, format='PNG')
            img_byte_array = img_byte_array.getvalue()
            st.download_button(label="処理された画像をダウンロード", data=img_byte_array, file_name='processed_image.png', mime='image/png')

    elif option == "画像をExcelに転記":
        st.subheader("画像をExcelに転記")
        st.write("画像をExcelシートに、1ピクセルが1セルに対応するように転記するアプリです。")

        image_file = st.file_uploader("画像ファイルをアップロードしてください", type=["jpg", "jpeg", "png"])

        if image_file is not None:
            image = Image.open(image_file)

            st.image(image, caption='Uploaded Image', use_column_width=True)

            output_excel_filename = st.text_input("出力Excelファイル名", "output_colors.xlsx")

            if st.button("変換してExcelに保存"):
                output_filepath = image_to_excel(image, output_excel_filename)
                st.success(f'画像の色情報をExcelに保存しました: {output_filepath}')

                download_button = st.download_button(
                    label="Excelファイルをダウンロード",
                    data=open(output_filepath, "rb").read(),
                    file_name=output_excel_filename,
                    key="download_button",
                )

                shutil.rmtree(os.path.dirname(output_filepath))

if __name__ == '__main__':
    main()
