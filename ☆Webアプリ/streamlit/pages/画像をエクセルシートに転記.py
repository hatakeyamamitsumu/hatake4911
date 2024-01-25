import streamlit as st
from PIL import Image
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill

def image_to_excel(image_stream, output_filename):
    # 画像ファイルを開く
    image = Image.open(image_stream)

    # 画像の幅と高さを取得
    width, height = image.size

    # 新しいExcelブックを作成
    workbook = Workbook()
    sheet = workbook.active

    # 画像の各ピクセルをExcelセルに対応させて色を設定
    for y in range(height):
        for x in range(width):
            pixel_color = image.getpixel((x, y))
            hex_color = f'{pixel_color[0]:02X}{pixel_color[1]:02X}{pixel_color[2]:02X}'  # RGBを16進数に変換
            fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')
            cell = sheet.cell(row=y + 1, column=x + 1)
            cell.fill = fill

    # Excelファイルを保存
    workbook.save(output_filename)
    return output_filename

# Streamlit UI
st.title("画像をエクセルに変換するアプリ")

uploaded_file = st.file_uploader("画像をアップロードしてください", type=["jpg", "jpeg", "png"])

if uploaded_file is not None:
    # アップロードされた画像を表示
    st.image(uploaded_file, caption="アップロードされた画像", use_column_width=True)

    # ボタンがクリックされたときに画像をExcelに変換
    if st.button("Excelに変換"):
        # 出力ファイル名を設定
        output_excel_filename = "画像をエクセル化.xlsx"
        
        # BytesIOオブジェクトを使って変換関数を呼び出す
        image_stream = BytesIO(uploaded_file.read())
        result_filename = image_to_excel(image_stream, output_excel_filename)

        # Excelファイルのダウンロードリンクを表示
        st.success(f"画像の色がExcelに保存されました: {result_filename}")
        st.download_button(
            label="Excelファイルをダウンロード",
            data=BytesIO(open(result_filename, "rb").read()),
            key="download_button",
            on_click=None,
            args=None,
            help="Excelファイルをダウンロードするにはクリックしてください。",
        )

