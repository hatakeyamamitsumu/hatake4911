import streamlit as st
from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import os
import shutil
import tempfile

def image_to_excel(image, output_filename):
    # 画像の幅と高さを取得
    width, height = image.size

    # 新しいExcelブックを作成
    workbook = Workbook()
    sheet = workbook.active

    # 画像の各ピクセルをExcelセルに対応させて色を設定
    for y in range(height):
        for x in range(width):
            pixel_color = image.getpixel((x, y))
            hex_color = f'{pixel_color[0]:02X}{pixel_color[1]:02X}{pixel_color[2]:02X}'  # RGBを16進数に変換
            fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')
            cell = sheet.cell(row=y + 1, column=x + 1)
            cell.fill = fill

    # 一時フォルダに保存
    temp_folder = tempfile.mkdtemp()
    temp_filepath = os.path.join(temp_folder, output_filename)
    workbook.save(temp_filepath)
    return temp_filepath

def main():
    st.title("画像を「セル情報として」エクセルに転記")
    st.text("大きな画像はとても時間がかかったり失敗したりするので、200KB以内の画像がおすすめです")

    # 画像ファイルをアップロード
    image_file = st.file_uploader("Upload Image File", type=["jpg", "jpeg", "png"])

    if image_file is not None:
        # 画像をPIL形式に変換
        image = Image.open(image_file)

        # 画像を表示
        st.image(image, caption='Uploaded Image', use_column_width=True)

        # 出力ファイル名を指定
        output_excel_filename = st.text_input("Output Excel Filename", "output_colors.xlsx")

        # 変換ボタン
        if st.button("Convert and Save to Excel"):
            output_filepath = image_to_excel(image, output_excel_filename)
            st.success(f'Image colors saved in Excel: {output_filepath}')

            # エクセルシートをダウンロードするボタン
            download_button = st.download_button(
                label="Download Excel File",
                data=open(output_filepath, "rb").read(),
                file_name=output_excel_filename,
                key="download_button",
            )

            # 一時フォルダをクリーンアップ
            shutil.rmtree(os.path.dirname(output_filepath))

if __name__ == "__main__":
    main()

